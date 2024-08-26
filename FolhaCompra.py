import pandas as pd
import openpyxl
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Border, Side
import numpy as np
import os


caminho = os.getcwd() 
arquivo_xls = caminho + r'\CACAU SHOW 201699-201700 TESTE.xls'
arquivo_xlsx = caminho + r'\CACAU_SHOW_201699-201700_TESTE_CONVERTIDO.xlsx'

# Ler o arquivo Excel original
xls = pd.read_excel(arquivo_xls, sheet_name=None)

# Converter e salvar em formato Excel
with pd.ExcelWriter(arquivo_xlsx, engine='openpyxl') as writer:
    for nome_planilha, df in xls.items():
        df.to_excel(writer, sheet_name=nome_planilha, index=False)

print("Conversão concluída!")

# Carregar a planilha convertida
Planilha_xml = pd.read_excel(arquivo_xlsx, sheet_name="COMPRA")

# Substituir valores vazios na coluna 'Cod. Produto' por NaN
Planilha_xml['Cod. Produto'] = Planilha_xml['Cod. Produto'].replace('', np.nan)

# Substituir NaN por strings vazias
Planilha_xml = Planilha_xml.fillna('')

# Salvar o DataFrame modificado de volta para o arquivo Excel
with pd.ExcelWriter(arquivo_xlsx, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    Planilha_xml.to_excel(writer, sheet_name="COMPRA", index=False)

print("Valores vazios na coluna 'Cod. Produto' substituídos por células em branco!")

# Adicionar a quebra de página e limpar células
wb = openpyxl.load_workbook(arquivo_xlsx)
ws = wb["COMPRA"]

# Adicionar quebras de página e limpar células onde 'Cod. Produto' está vazio
for idx, row in Planilha_xml.iterrows():
    if row['Cod. Produto'] == '':
        # Limpar células nas colunas 'Cte' e 'Soma de Qtde'
        ws[f'A{idx + 2}'].value = None
        ws[f'F{idx + 2}'].value = None
        # Adicionar quebra de página após a linha onde 'Cod. Produto' está vazio
        ws.row_breaks.append(Break(id=idx + 2))  # +2 para adicionar a quebra após a linha com 'Cod. Produto' vazio

# Ajustar a largura das colunas específicas
colunas_para_ajustar = ['Cod. Produto', 'ALOCAÇÃO', 'Soma de Qtde']

for coluna in colunas_para_ajustar:
    col_index = Planilha_xml.columns.get_loc(coluna) + 1  # +1 porque openpyxl é 1-based index
    col_letter = openpyxl.utils.get_column_letter(col_index)

    max_length = 0
    for cell in ws[col_letter]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass

    # Definir a largura da coluna
    ws.column_dimensions[col_letter].width = max_length + 2  # +2 para algum espaçamento extra

# Ajustar a largura da coluna 'Descrição'
descricao_col = 'Descrição'
descricao_index = Planilha_xml.columns.get_loc(descricao_col) + 1
descricao_letter = openpyxl.utils.get_column_letter(descricao_index)

descricao_max_length = 0
for cell in ws[descricao_letter]:
    try:
        if len(str(cell.value)) > descricao_max_length:
            descricao_max_length = len(cell.value)
    except:
        pass



# Definir a largura da coluna 'Descrição'
ws.column_dimensions[descricao_letter].width = descricao_max_length + 2  # +2 para algum espaçamento extra

# Adicionar bordas horizontais às linhas
border_top = Side(style='thin')
border_bottom = Side(style='thin')
border_horizontal = Border(top=border_top, bottom=border_bottom)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Começa a partir da segunda linha para evitar o cabeçalho
    for cell in row:
        cell.border = border_horizontal

# Configurar a orientação da página para paisagem
ws.page_setup.orientation = 'landscape'


# Ajustar a altura das linhas
row_height = 22  # Defina a altura desejada para as linhas
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = row_height


# Salvar o arquivo Excel com as quebras de página e células limpas
wb.save(arquivo_xlsx)

print("Quebras de página adicionadas, células limpas, larguras das colunas ajustadas, e bordas horizontais adicionadas às linhas!")
print('Terminou :)')