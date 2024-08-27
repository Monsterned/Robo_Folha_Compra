import pandas as pd
import openpyxl
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Border, Side
import numpy as np
import os
from openpyxl.styles import Alignment

caminho = os.getcwd() 

# Carregar os DataFrames
BASE = pd.read_excel('CACAU SHOW 202191 RESUMIDA.xlsx', sheet_name='BASE')
ALOCAÇÃO = pd.read_excel('CACAU SHOW 202191 RESUMIDA.xlsx', sheet_name='ALOCAÇÃO')
ROTAS = pd.read_excel('ROTAS TERMINAL CACAU.xlsx', sheet_name='Plan1', usecols="A:B")

# Verificar as colunas disponíveis
print("Colunas em BASE:", BASE.columns)
print("Colunas em ALOCAÇÃO:", ALOCAÇÃO.columns)
print("Colunas em ROTAS:", ROTAS.columns)

# Realizar o merge para buscar a coluna desejada em ALOCAÇÃO
resultado = pd.merge(BASE, ALOCAÇÃO.iloc[:, [0, 3]], how='left', left_on='Cod. Produto', right_on=ALOCAÇÃO.columns[0])
BASE['ALOCAÇÃO1'] = resultado.iloc[:, -1]

# Criar a nova coluna 'ALOCAÇÃO2' com os três primeiros caracteres da coluna 'ALOCAÇÃO1'
BASE['ALOCAÇÃO2'] = BASE['ALOCAÇÃO1'].str[:3]

# Ajustar o nome da coluna de pesquisa conforme necessário
BASE = pd.merge(BASE, ROTAS, how='left', left_on='ALOCAÇÃO2', right_on=ROTAS.columns[0])

# Renomear a coluna resultante para 'ROTA' (ou o nome desejado)
BASE.rename(columns={ROTAS.columns[1]: 'ROTA'}, inplace=True)

# Remover a coluna de pesquisa se não for mais necessária
BASE.drop(columns=[ROTAS.columns[0]], inplace=True)

# Salvar o DataFrame atualizado de volta em um arquivo Excel na guia BASE
with pd.ExcelWriter('CACAU SHOW 202191 RESUMIDA.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    BASE.to_excel(writer, sheet_name='BASE', index=False)

# Carregar os DataFrames
BASE = pd.read_excel('CACAU SHOW 202191 RESUMIDA.xlsx', sheet_name='BASE')
ALOCAÇÃO = pd.read_excel('CACAU SHOW 202191 RESUMIDA.xlsx', sheet_name='ALOCAÇÃO')
ROTAS = pd.read_excel('ROTAS TERMINAL CACAU.xlsx', sheet_name='Plan1', usecols="A:B")

# Realizar o merge para buscar a coluna desejada em ALOCAÇÃO
resultado = pd.merge(BASE, ALOCAÇÃO.iloc[:, [0, 3]], how='left', left_on='Cod. Produto', right_on=ALOCAÇÃO.columns[0])
BASE['ALOCAÇÃO1'] = resultado.iloc[:, -1]

# Criar a nova coluna 'ALOCAÇÃO2' com os três primeiros caracteres da coluna 'ALOCAÇÃO1'
BASE['ALOCAÇÃO2'] = BASE['ALOCAÇÃO1'].str[:3]

# Realizar o merge com o DataFrame ROTAS para adicionar a coluna desejada
BASE = pd.merge(BASE, ROTAS, how='left', left_on='ALOCAÇÃO2', right_on=ROTAS.columns[0])

# Renomear a coluna resultante para 'ROTA'
BASE.rename(columns={ROTAS.columns[1]: 'ROTA'}, inplace=True)

# Remover a coluna de pesquisa se não for mais necessária
BASE.drop(columns=[ROTAS.columns[0]], inplace=True)

# Criar a Tabela Dinâmica
tabela_dinamica = BASE.pivot_table(
    index=['Cte', 'Cidade', 'Descrição', 'Cod. Produto', 'ALOCAÇÃO2'],
    values='Qtde',
    aggfunc='sum',
    fill_value=0
).reset_index()

# Ajustar o tipo da coluna 'Cte' para string
tabela_dinamica['Cte'] = tabela_dinamica['Cte'].astype(str)

# Adicionar totais antes de cada novo Cte
def add_totals(df):
    result = []
    for cte in df['Cte'].unique():
        subset = df[df['Cte'] == cte]
        total_row = subset[['Qtde']].sum()
        total_row = total_row.to_frame().T
        total_row['Cte'] = ''
        total_row['Cidade'] = ''
        total_row['Descrição'] = ''
        total_row['Cod. Produto'] = ''
        total_row['ALOCAÇÃO2'] = ''
        result.append(subset)
        result.append(total_row)
    return pd.concat(result, ignore_index=True)

# Adicionar totais
tabela_dinamica = add_totals(tabela_dinamica)

# Salvar o DataFrame atualizado e a Tabela Dinâmica de volta em um arquivo Excel
with pd.ExcelWriter('CACAU SHOW 202191 RESUMIDA.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    BASE.to_excel(writer, sheet_name='BASE', index=False)
    tabela_dinamica.to_excel(writer, sheet_name='DINAMICA', index=False)

arquivo_xlsx = caminho + r'\CACAU SHOW 202191 RESUMIDA.xlsx'

# Carregar a planilha convertida
Planilha_xml = pd.read_excel(arquivo_xlsx, sheet_name="DINAMICA")
Planilha_xml = Planilha_xml.rename(columns={'ALOCAÇÃO2': 'ALOCAÇÃO'})
# Substituir valores vazios na coluna 'Cod. Produto' por NaN
Planilha_xml['Cod. Produto'] = Planilha_xml['Cod. Produto'].replace('', np.nan)

# Substituir NaN por strings vazias
Planilha_xml = Planilha_xml.fillna('')

# Salvar o DataFrame modificado de volta para o arquivo Excel
with pd.ExcelWriter(arquivo_xlsx, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    Planilha_xml.to_excel(writer, sheet_name="DINAMICA", index=False)

print("Valores vazios na coluna 'Cod. Produto' substituídos por células em branco!")

# Adicionar a quebra de página e limpar células
wb = openpyxl.load_workbook(arquivo_xlsx)
ws = wb["DINAMICA"]

# Adicionar quebras de página e limpar células onde 'Cod. Produto' está vazio
for idx, row in Planilha_xml.iterrows():
    if row['Cod. Produto'] == '':
        # Limpar células nas colunas 'Cte' e 'Soma de Qtde'
        ws[f'A{idx + 2}'].value = None
        ws[f'F{idx + 2}'].value = None
        # Adicionar quebra de página após a linha onde 'Cod. Produto' está vazio
        ws.row_breaks.append(Break(id=idx + 2))  # +2 para adicionar a quebra após a linha com 'Cod. Produto' vazio

# Ajustar a largura das colunas específicas
colunas_para_ajustar = ['Cod. Produto', 'ALOCAÇÃO', 'Qtde']

for coluna in colunas_para_ajustar:
    col_index = Planilha_xml.columns.get_loc(coluna) + 1  # +1 porque openpyxl é 1-based index
    col_letter = openpyxl.utils.get_column_letter(col_index)

    max_length = 0
    for cell in ws[col_letter]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass

    # Definir a largura da coluna
    ws.column_dimensions[col_letter].width = max_length + 2  # +2 para algum espaçamento extra

# Ajustar a largura da coluna 'Descrição'
descricao_col = 'Descrição'
descricao_index = Planilha_xml.columns.get_loc(descricao_col) + 1
descricao_letter = openpyxl.utils.get_column_letter(descricao_index)

descricao_max_length = 0
for cell in ws[descricao_letter]:
    try:
        if len(str(cell.value)) > descricao_max_length:
            descricao_max_length = len(cell.value)
    except:
        pass

# Definir a largura da coluna 'Descrição'
ws.column_dimensions[descricao_letter].width = descricao_max_length + 2  # +2 para algum espaçamento extra


# Adicionar bordas horizontais às linhas
border_top = Side(style='thin')
border_bottom = Side(style='thin')
border_horizontal = Border(top=border_top, bottom=border_bottom)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Começa a partir da segunda linha para evitar o cabeçalho
    for cell in row:
        cell.border = border_horizontal

# Configurar a orientação da página para paisagem
ws.page_setup.orientation = 'landscape'


# Ajustar a altura das linhas
row_height = 22  # Defina a altura desejada para as linhas
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = row_height

# Salvar o arquivo Excel com as quebras de página e células limpas
wb.save(arquivo_xlsx)

print("Quebras de página adicionadas, células limpas, larguras das colunas ajustadas, e bordas horizontais adicionadas às linhas!")
print('Terminou :)')